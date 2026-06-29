"""
app/backup.py — Excel-inside-ZIP backup with debounced triggering.

Public interface:
    trigger_backup(reason)   — call after any data change; debounces to one run per 2 s
"""
import io
import os
import threading
import zipfile
from datetime import datetime

from app.config import UPLOAD_FOLDER
from app.db import get_db
from app.utils import now


# ---------------------------------------------------------------------------
# Settings helpers (imported lazily to avoid circular imports)
# ---------------------------------------------------------------------------

def _load_settings():
    from app.routes.settings import load_settings
    return load_settings()


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

def _build_excel(conn) -> bytes:
    """Return a patients.xlsx workbook as raw bytes."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill("solid", fgColor="1A3A2A")
    ALT_FILL    = PatternFill("solid", fgColor="F0F7F4")

    def style_header(ws, row_num, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row_num].height = 22

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def alt_rows(ws, start, end, ncols):
        for r in range(start, end + 1):
            if r % 2 == 0:
                for c in range(1, ncols + 1):
                    ws.cell(row=r, column=c).fill = ALT_FILL

    wb = openpyxl.Workbook()

    # ── Sheet 1: All Patients ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "All Patients"

    patients = conn.execute("""
        SELECT p.*, COUNT(mh.id) as total_visits
        FROM patients p
        LEFT JOIN medical_history mh ON p.id = mh.patient_id
        GROUP BY p.id ORDER BY p.patient_id ASC
    """).fetchall()

    ws1.merge_cells("A1:M1")
    tc = ws1["A1"]
    tc.value     = f"MediRecord — Patient Backup   |   {now()}"
    tc.font      = Font(bold=True, size=13, color="1A3A2A")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    p_headers = [
        "#", "Patient ID", "Full Name", "Age", "Gender", "Blood Group",
        "Phone", "Email", "Address", "Emergency Contact",
        "Total Visits", "Registered On", "Last Updated",
    ]
    for col, h in enumerate(p_headers, 1):
        ws1.cell(row=2, column=col, value=h)
    style_header(ws1, 2, len(p_headers))

    for i, p in enumerate(patients, 1):
        r = i + 2
        row_data = [
            i, p["patient_id"], p["name"], p["age"] or "",
            p["gender"] or "", p["blood_group"] or "",
            p["phone"] or "", p["email"] or "", p["address"] or "",
            p["emergency_contact"] or "", p["total_visits"],
            p["created_at"], p["updated_at"],
        ]
        for col, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=col, value=val)

    alt_rows(ws1, 3, len(patients) + 2, len(p_headers))
    set_col_widths(ws1, [4, 10, 24, 6, 10, 12, 16, 24, 30, 24, 10, 18, 18])
    ws1.freeze_panes = "A3"

    # ── Sheet 2: Medical History ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Medical History")

    history = conn.execute("""
        SELECT p.patient_id, p.name as patient_name, p.age, p.gender,
               p.blood_group, p.phone,
               mh.visit_date, mh.disease, mh.diagnosis,
               mh.prescription, mh.notes, mh.doctor_name, mh.created_at
        FROM medical_history mh
        JOIN patients p ON p.id = mh.patient_id
        ORDER BY p.patient_id ASC, mh.visit_date DESC
    """).fetchall()

    ws2.merge_cells("A1:N1")
    tc2 = ws2["A1"]
    tc2.value     = f"MediRecord — Medical History Backup   |   {now()}"
    tc2.font      = Font(bold=True, size=13, color="1A3A2A")
    tc2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    h_headers = [
        "#", "Patient ID", "Patient Name", "Age", "Gender", "Blood Group", "Phone",
        "Visit Date", "Disease / Complaint", "Diagnosis",
        "Prescription", "Notes", "Doctor Name", "Record Created",
    ]
    for col, h in enumerate(h_headers, 1):
        ws2.cell(row=2, column=col, value=h)
    style_header(ws2, 2, len(h_headers))

    for i, h in enumerate(history, 1):
        r = i + 2
        row_data = [
            i, h["patient_id"], h["patient_name"], h["age"] or "",
            h["gender"] or "", h["blood_group"] or "", h["phone"] or "",
            h["visit_date"], h["disease"] or "", h["diagnosis"] or "",
            h["prescription"] or "", h["notes"] or "",
            h["doctor_name"] or "", h["created_at"],
        ]
        for col, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=col, value=val)

    alt_rows(ws2, 3, len(history) + 2, len(h_headers))
    set_col_widths(ws2, [4, 10, 22, 6, 10, 12, 16, 12, 22, 28, 28, 28, 18, 18])
    ws2.freeze_panes = "A3"

    # ── Sheet 3: Uploaded Files Index ────────────────────────────────────────
    ws3 = wb.create_sheet("Uploaded Files")

    files = conn.execute("""
        SELECT p.patient_id, p.name as patient_name,
               r.original_name, r.report_type, r.description, r.uploaded_at, r.filename
        FROM reports r
        JOIN patients p ON p.id = r.patient_id
        ORDER BY p.patient_id ASC, r.uploaded_at DESC
    """).fetchall()

    ws3.merge_cells("A1:G1")
    tc3 = ws3["A1"]
    tc3.value     = f"MediRecord — Uploaded Files Index   |   {now()}"
    tc3.font      = Font(bold=True, size=13, color="1A3A2A")
    tc3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    f_headers = ["#", "Patient ID", "Patient Name", "File Name", "Report Type", "Description", "Uploaded On"]
    for col, h in enumerate(f_headers, 1):
        ws3.cell(row=2, column=col, value=h)
    style_header(ws3, 2, len(f_headers))

    for i, f in enumerate(files, 1):
        r = i + 2
        row_data = [
            i, f["patient_id"], f["patient_name"], f["original_name"],
            f["report_type"] or "", f["description"] or "", f["uploaded_at"],
        ]
        for col, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=col, value=val)

    alt_rows(ws3, 3, len(files) + 2, len(f_headers))
    set_col_widths(ws3, [4, 10, 22, 28, 16, 28, 18])
    ws3.freeze_panes = "A3"

    # ── Sheet 4: Summary ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Summary")
    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 40

    total_file_size = 0
    for f in files:
        fp = os.path.join(UPLOAD_FOLDER, f["patient_id"], f["filename"])
        if os.path.exists(fp):
            total_file_size += os.path.getsize(fp)

    summary = [
        ("MediRecord Backup Summary", ""),
        ("", ""),
        ("Generated On",    now()),
        ("Total Patients",  len(patients)),
        ("Total Visits",    len(history)),
        ("Total Files",     len(files)),
        ("Total File Size", f"{total_file_size / (1024 * 1024):.2f} MB"),
    ]
    for i, (label, value) in enumerate(summary, 1):
        ws4.cell(row=i, column=1, value=label)
        ws4.cell(row=i, column=2, value=str(value))
        if i == 1:
            ws4.cell(row=i, column=1).font = Font(bold=True, size=14, color="1A3A2A")
        elif label:
            ws4.cell(row=i, column=1).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Backup entry point
# ---------------------------------------------------------------------------

def create_backup(trigger: str = "change") -> None:
    """
    Build an Excel+uploads ZIP in the configured backup folder.
    Safe to call from a background thread.
    """
    settings      = _load_settings()
    backup_folder = settings.get("backup_folder", "").strip()

    if not backup_folder:
        print(f"[Backup] Skipped ({trigger}) — no backup folder set.")
        return

    try:
        os.makedirs(backup_folder, exist_ok=True)
    except Exception as e:
        print(f"[Backup] Cannot create folder: {e}")
        return

    date_str    = datetime.now().strftime("%Y-%m-%d")
    backup_file = os.path.join(backup_folder, f"MediRecord_Backup_{date_str}.zip")

    try:
        conn = get_db()
        try:
            excel_bytes = _build_excel(conn)
        finally:
            conn.close()

        files_added = 0
        with zipfile.ZipFile(backup_file, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("patients.xlsx", excel_bytes)

            if os.path.exists(UPLOAD_FOLDER):
                for pid_folder in sorted(os.listdir(UPLOAD_FOLDER)):
                    pid_path = os.path.join(UPLOAD_FOLDER, pid_folder)
                    if os.path.isdir(pid_path) and pid_folder.startswith("P-"):
                        for fname in sorted(os.listdir(pid_path)):
                            fpath = os.path.join(pid_path, fname)
                            if os.path.isfile(fpath):
                                zf.write(fpath, f"uploads/{pid_folder}/{fname}")
                                files_added += 1

        size_mb = os.path.getsize(backup_file) / (1024 * 1024)
        print(f"[Backup] OK ({trigger}) {files_added} files → {backup_file} ({size_mb:.2f} MB)")

    except ImportError:
        print("[Backup] ERROR: openpyxl not installed. Run: pip install openpyxl")
    except Exception as e:
        print(f"[Backup] ERROR: {e}")


# ---------------------------------------------------------------------------
# Debounced trigger
# ---------------------------------------------------------------------------

_backup_timer: threading.Timer | None = None
_backup_lock  = threading.Lock()


def trigger_backup(reason: str = "change") -> None:
    """
    Schedule create_backup to run 2 s after the last call.
    Rapid back-to-back changes collapse into a single backup run.
    """
    global _backup_timer
    with _backup_lock:
        if _backup_timer is not None:
            _backup_timer.cancel()
        t = threading.Timer(2.0, create_backup, args=(reason,))
        t.daemon = True
        t.start()
        _backup_timer = t
