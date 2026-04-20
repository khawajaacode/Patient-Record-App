import os
import sys
import sqlite3
import uuid
import threading
import webbrowser
import json
import zipfile
import shutil
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_from_directory, abort
from werkzeug.utils import secure_filename

# ──────────────────────────────────────────────
# PATH RESOLUTION
# ──────────────────────────────────────────────

def get_base_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))

def get_resource_dir():
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))

BASE_DIR     = get_base_dir()
RESOURCE_DIR = get_resource_dir()

# ──────────────────────────────────────────────
# FLASK SETUP
# ──────────────────────────────────────────────

app = Flask(__name__, template_folder=os.path.join(RESOURCE_DIR, 'templates'))
app.config['UPLOAD_FOLDER']      = os.path.join(BASE_DIR, 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
app.config['SECRET_KEY']         = 'local-medirecord-offline-key'
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif', 'bmp', 'tiff'}
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# ──────────────────────────────────────────────
# SETTINGS
# ──────────────────────────────────────────────

SETTINGS_FILE = os.path.join(BASE_DIR, 'settings.json')

def load_settings():
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r') as f:
                return json.load(f)
        except Exception:
            pass
    return {'backup_folder': ''}

def save_settings(s):
    try:
        with open(SETTINGS_FILE, 'w') as f:
            json.dump(s, f, indent=2)
    except Exception as e:
        print(f"[Settings] Save failed: {e}")

# ──────────────────────────────────────────────
# DATABASE
# ──────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(os.path.join(BASE_DIR, 'patients.db'))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()

    # patients — patient_id is short human-readable (P-0001), id is UUID primary key
    c.execute('''CREATE TABLE IF NOT EXISTS patients (
        id           TEXT PRIMARY KEY,
        patient_id   TEXT UNIQUE NOT NULL,
        name         TEXT NOT NULL,
        age          INTEGER,
        gender       TEXT,
        blood_group  TEXT,
        phone        TEXT,
        email        TEXT,
        address      TEXT,
        emergency_contact TEXT,
        created_at   TEXT NOT NULL,
        updated_at   TEXT NOT NULL
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS medical_history (
        id          TEXT PRIMARY KEY,
        patient_id  TEXT NOT NULL,
        visit_date  TEXT NOT NULL,
        disease     TEXT,
        diagnosis   TEXT,
        prescription TEXT,
        notes       TEXT,
        doctor_name TEXT,
        created_at  TEXT NOT NULL,
        FOREIGN KEY (patient_id) REFERENCES patients(id) ON DELETE CASCADE
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS reports (
        id            TEXT PRIMARY KEY,
        patient_id    TEXT NOT NULL,
        history_id    TEXT,
        filename      TEXT NOT NULL,
        original_name TEXT NOT NULL,
        file_type     TEXT NOT NULL,
        report_type   TEXT,
        description   TEXT,
        uploaded_at   TEXT NOT NULL,
        FOREIGN KEY (patient_id) REFERENCES patients(id) ON DELETE CASCADE
    )''')

    # counter table for generating sequential patient IDs
    c.execute('''CREATE TABLE IF NOT EXISTS id_counter (
        name  TEXT PRIMARY KEY,
        value INTEGER NOT NULL DEFAULT 0
    )''')
    c.execute("INSERT OR IGNORE INTO id_counter (name, value) VALUES ('patient', 0)")

    conn.commit()
    conn.close()

# ──────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────

def allowed_file(f):
    return '.' in f and f.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def now():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def uid():
    return str(uuid.uuid4())

def generate_patient_id(conn):
    """Generate next sequential patient ID like P-0001, P-0002 …"""
    conn.execute("UPDATE id_counter SET value = value + 1 WHERE name = 'patient'")
    row = conn.execute("SELECT value FROM id_counter WHERE name = 'patient'").fetchone()
    return f"P-{row['value']:04d}"

def patient_upload_dir(short_id):
    """Returns and creates /uploads/P-0001/ folder."""
    folder = os.path.join(app.config['UPLOAD_FOLDER'], short_id)
    os.makedirs(folder, exist_ok=True)
    return folder

# ──────────────────────────────────────────────
# BACKUP — ZIP with Excel + all uploaded files
# ──────────────────────────────────────────────

def create_backup(trigger='change'):
    """
    Creates MediRecord_Backup_YYYY-MM-DD.zip in the doctor's chosen folder.
    ZIP contains:
      patients.xlsx  — all patient data + medical history
      uploads/       — all uploaded files in P-XXXX/ subfolders
    Always overwrites same-named file for today's date.
    """
    settings      = load_settings()
    backup_folder = settings.get('backup_folder', '').strip()

    if not backup_folder:
        print(f"[Backup] Skipped ({trigger}) — no backup folder set.")
        return

    try:
        os.makedirs(backup_folder, exist_ok=True)
    except Exception as e:
        print(f"[Backup] Cannot create folder: {e}")
        return

    date_str    = datetime.now().strftime('%Y-%m-%d')
    backup_file = os.path.join(backup_folder, f'MediRecord_Backup_{date_str}.zip')
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
        HEADER_FILL = PatternFill('solid', fgColor='1A3A2A')
        ALT_FILL    = PatternFill('solid', fgColor='F0F7F4')

        def style_header(ws, row_num, num_cols):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.font      = HEADER_FONT
                cell.fill      = HEADER_FILL
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.row_dimensions[row_num].height = 22

        def set_col_widths(ws, widths):
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        def alt_rows(ws, start, end, ncols):
            for r in range(start, end + 1):
                if r % 2 == 0:
                    for c in range(1, ncols + 1):
                        ws.cell(row=r, column=c).fill = ALT_FILL

        conn = get_db()
        wb   = openpyxl.Workbook()

        # ── Sheet 1: All Patients ────────────────────────────────
        ws1 = wb.active
        ws1.title = 'All Patients'

        ws1.merge_cells('A1:M1')
        tc = ws1['A1']
        tc.value     = f'MediRecord — Patient Backup   |   {now()}   |   Trigger: {trigger}'
        tc.font      = Font(bold=True, size=13, color='1A3A2A')
        tc.alignment = Alignment(horizontal='center', vertical='center')
        ws1.row_dimensions[1].height = 28

        p_headers = [
            '#', 'Patient ID', 'Full Name', 'Age', 'Gender', 'Blood Group',
            'Phone', 'Email', 'Address', 'Emergency Contact',
            'Total Visits', 'Registered On', 'Last Updated'
        ]
        for col, h in enumerate(p_headers, 1):
            ws1.cell(row=2, column=col, value=h)
        style_header(ws1, 2, len(p_headers))

        patients = conn.execute('''
            SELECT p.*, COUNT(mh.id) as total_visits
            FROM patients p
            LEFT JOIN medical_history mh ON p.id = mh.patient_id
            GROUP BY p.id ORDER BY p.patient_id ASC
        ''').fetchall()

        for i, p in enumerate(patients, 1):
            r = i + 2
            ws1.cell(row=r, column=1,  value=i)
            ws1.cell(row=r, column=2,  value=p['patient_id'])
            ws1.cell(row=r, column=3,  value=p['name'])
            ws1.cell(row=r, column=4,  value=p['age'] or '')
            ws1.cell(row=r, column=5,  value=p['gender'] or '')
            ws1.cell(row=r, column=6,  value=p['blood_group'] or '')
            ws1.cell(row=r, column=7,  value=p['phone'] or '')
            ws1.cell(row=r, column=8,  value=p['email'] or '')
            ws1.cell(row=r, column=9,  value=p['address'] or '')
            ws1.cell(row=r, column=10, value=p['emergency_contact'] or '')
            ws1.cell(row=r, column=11, value=p['total_visits'])
            ws1.cell(row=r, column=12, value=p['created_at'])
            ws1.cell(row=r, column=13, value=p['updated_at'])

        alt_rows(ws1, 3, len(patients) + 2, len(p_headers))
        set_col_widths(ws1, [4, 10, 24, 6, 10, 12, 16, 24, 30, 24, 10, 18, 18])
        ws1.freeze_panes = 'A3'

        # ── Sheet 2: Medical History ─────────────────────────────
        ws2 = wb.create_sheet('Medical History')

        ws2.merge_cells('A1:N1')
        tc2 = ws2['A1']
        tc2.value     = f'MediRecord — Medical History Backup   |   {now()}'
        tc2.font      = Font(bold=True, size=13, color='1A3A2A')
        tc2.alignment = Alignment(horizontal='center', vertical='center')
        ws2.row_dimensions[1].height = 28

        h_headers = [
            '#', 'Patient ID', 'Patient Name', 'Age', 'Gender', 'Blood Group', 'Phone',
            'Visit Date', 'Disease / Complaint', 'Diagnosis',
            'Prescription', 'Notes', 'Doctor Name', 'Record Created'
        ]
        for col, h in enumerate(h_headers, 1):
            ws2.cell(row=2, column=col, value=h)
        style_header(ws2, 2, len(h_headers))

        history = conn.execute('''
            SELECT p.patient_id, p.name as patient_name, p.age, p.gender,
                   p.blood_group, p.phone,
                   mh.visit_date, mh.disease, mh.diagnosis,
                   mh.prescription, mh.notes, mh.doctor_name, mh.created_at
            FROM medical_history mh
            JOIN patients p ON p.id = mh.patient_id
            ORDER BY p.patient_id ASC, mh.visit_date DESC
        ''').fetchall()

        for i, h in enumerate(history, 1):
            r = i + 2
            ws2.cell(row=r, column=1,  value=i)
            ws2.cell(row=r, column=2,  value=h['patient_id'])
            ws2.cell(row=r, column=3,  value=h['patient_name'])
            ws2.cell(row=r, column=4,  value=h['age'] or '')
            ws2.cell(row=r, column=5,  value=h['gender'] or '')
            ws2.cell(row=r, column=6,  value=h['blood_group'] or '')
            ws2.cell(row=r, column=7,  value=h['phone'] or '')
            ws2.cell(row=r, column=8,  value=h['visit_date'])
            ws2.cell(row=r, column=9,  value=h['disease'] or '')
            ws2.cell(row=r, column=10, value=h['diagnosis'] or '')
            ws2.cell(row=r, column=11, value=h['prescription'] or '')
            ws2.cell(row=r, column=12, value=h['notes'] or '')
            ws2.cell(row=r, column=13, value=h['doctor_name'] or '')
            ws2.cell(row=r, column=14, value=h['created_at'])

        alt_rows(ws2, 3, len(history) + 2, len(h_headers))
        set_col_widths(ws2, [4, 10, 22, 6, 10, 12, 16, 12, 22, 28, 28, 28, 18, 18])
        ws2.freeze_panes = 'A3'

        # ── Sheet 3: Uploaded Files Index ────────────────────────
        ws3 = wb.create_sheet('Uploaded Files')

        ws3.merge_cells('A1:G1')
        tc3 = ws3['A1']
        tc3.value     = f'MediRecord — Uploaded Files Index   |   {now()}'
        tc3.font      = Font(bold=True, size=13, color='1A3A2A')
        tc3.alignment = Alignment(horizontal='center', vertical='center')
        ws3.row_dimensions[1].height = 28

        f_headers = ['#', 'Patient ID', 'Patient Name', 'File Name', 'Report Type', 'Description', 'Uploaded On']
        for col, h in enumerate(f_headers, 1):
            ws3.cell(row=2, column=col, value=h)
        style_header(ws3, 2, len(f_headers))

        files = conn.execute('''
            SELECT p.patient_id, p.name as patient_name,
                   r.original_name, r.report_type, r.description, r.uploaded_at, r.filename
            FROM reports r
            JOIN patients p ON p.id = r.patient_id
            ORDER BY p.patient_id ASC, r.uploaded_at DESC
        ''').fetchall()

        for i, f in enumerate(files, 1):
            r = i + 2
            ws3.cell(row=r, column=1, value=i)
            ws3.cell(row=r, column=2, value=f['patient_id'])
            ws3.cell(row=r, column=3, value=f['patient_name'])
            ws3.cell(row=r, column=4, value=f['original_name'])
            ws3.cell(row=r, column=5, value=f['report_type'] or '')
            ws3.cell(row=r, column=6, value=f['description'] or '')
            ws3.cell(row=r, column=7, value=f['uploaded_at'])

        alt_rows(ws3, 3, len(files) + 2, len(f_headers))
        set_col_widths(ws3, [4, 10, 22, 28, 16, 28, 18])
        ws3.freeze_panes = 'A3'

        # ── Sheet 4: Summary ─────────────────────────────────────
        ws4 = wb.create_sheet('Summary')
        ws4.column_dimensions['A'].width = 28
        ws4.column_dimensions['B'].width = 40

        total_file_size = 0
        for f in files:
            fp = os.path.join(app.config['UPLOAD_FOLDER'], f['patient_id'], f['filename'])
            if os.path.exists(fp):
                total_file_size += os.path.getsize(fp)

        summary = [
            ('MediRecord Backup Summary', ''),
            ('', ''),
            ('Generated On',    now()),
            ('Triggered By',    trigger),
            ('Backup Date',     date_str),
            ('', ''),
            ('Total Patients',  len(patients)),
            ('Total Visits',    len(history)),
            ('Total Files',     len(files)),
            ('Total File Size', f"{total_file_size / (1024*1024):.2f} MB"),
            ('', ''),
            ('Backup Location', backup_file),
        ]
        for i, (label, value) in enumerate(summary, 1):
            ws4.cell(row=i, column=1, value=label)
            ws4.cell(row=i, column=2, value=str(value))
            if i == 1:
                ws4.cell(row=i, column=1).font = Font(bold=True, size=14, color='1A3A2A')
            elif label:
                ws4.cell(row=i, column=1).font = Font(bold=True)

        conn.close()

        # Save Excel into in-memory bytes (no temp file — avoids race condition)
        import io as _io
        excel_buf = _io.BytesIO()
        wb.save(excel_buf)
        excel_buf.seek(0)
        excel_bytes = excel_buf.read()

        # ── Build ZIP ────────────────────────────────────────────
        with zipfile.ZipFile(backup_file, 'w', zipfile.ZIP_DEFLATED) as zf:
            # Add Excel from memory
            zf.writestr('patients.xlsx', excel_bytes)

            # Add all uploaded files maintaining P-XXXX folder structure
            uploads_root = app.config['UPLOAD_FOLDER']
            files_added  = 0
            if os.path.exists(uploads_root):
                for pid_folder in sorted(os.listdir(uploads_root)):
                    pid_path = os.path.join(uploads_root, pid_folder)
                    if os.path.isdir(pid_path) and pid_folder.startswith('P-'):
                        for fname in sorted(os.listdir(pid_path)):
                            fpath = os.path.join(pid_path, fname)
                            if os.path.isfile(fpath):
                                # Inside zip: uploads/P-0001/xray.pdf
                                arcname = 'uploads/' + pid_folder + '/' + fname
                                zf.write(fpath, arcname)
                                files_added += 1

        size_mb = os.path.getsize(backup_file) / (1024 * 1024)
        print(f"[Backup] OK ({trigger}) {len(patients)} patients, {len(history)} visits, "
              f"{files_added} files -> {backup_file} ({size_mb:.2f} MB)")

    except ImportError:
        print("[Backup] ERROR: openpyxl not installed. Run: pip install openpyxl")
    except Exception as e:
        print(f"[Backup] ERROR: {e}")

# Debounce: waits 2s after the last trigger, then runs once.
# Rapid changes (add patient + upload file) collapse into ONE backup.
# Nothing is ever silently dropped — backup ALWAYS runs eventually.
_backup_timer = None
_backup_lock  = threading.Lock()

def trigger_backup(reason='change'):
    global _backup_timer
    with _backup_lock:
        if _backup_timer is not None:
            _backup_timer.cancel()
        t = threading.Timer(2.0, create_backup, args=(reason,))
        t.daemon = True
        t.start()
        _backup_timer = t

# ──────────────────────────────────────────────
# PAGES
# ──────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/patient/<patient_id>')
def patient_detail(patient_id):
    return render_template('patient_detail.html', patient_id=patient_id)

# ──────────────────────────────────────────────
# API — SETTINGS
# ──────────────────────────────────────────────

@app.route('/api/settings', methods=['GET'])
def get_settings_api():
    return jsonify(load_settings())

@app.route('/api/settings', methods=['POST'])
def save_settings_api():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data'}), 400
    folder = data.get('backup_folder', '').strip()
    if folder:
        try:
            os.makedirs(folder, exist_ok=True)
        except Exception as e:
            return jsonify({'error': f'Cannot create folder: {e}'}), 400
    s = load_settings()
    s['backup_folder'] = folder
    save_settings(s)
    trigger_backup('folder_set')
    return jsonify({'message': 'Saved', 'backup_folder': folder})

@app.route('/api/backup/now', methods=['POST'])
def manual_backup():
    if not load_settings().get('backup_folder', '').strip():
        return jsonify({'error': 'No backup folder set. Open ⚙ Settings first.'}), 400
    trigger_backup('manual')
    return jsonify({'message': 'Backup started in background.'})

# ──────────────────────────────────────────────
# API — PATIENTS
# ──────────────────────────────────────────────

@app.route('/api/patients', methods=['GET'])
def get_patients():
    s = request.args.get('search', '').strip()
    conn = get_db()
    try:
        if s:
            rows = conn.execute('''
                SELECT p.*, COUNT(mh.id) as visit_count FROM patients p
                LEFT JOIN medical_history mh ON p.id = mh.patient_id
                WHERE p.name LIKE ? OR p.phone LIKE ? OR p.blood_group LIKE ?
                   OR p.patient_id LIKE ?
                   OR EXISTS (SELECT 1 FROM medical_history m2 WHERE m2.patient_id = p.id
                              AND (m2.disease LIKE ? OR m2.diagnosis LIKE ?))
                GROUP BY p.id ORDER BY p.patient_id ASC
            ''', (f'%{s}%', f'%{s}%', f'%{s}%', f'%{s}%', f'%{s}%', f'%{s}%')).fetchall()
        else:
            rows = conn.execute('''
                SELECT p.*, COUNT(mh.id) as visit_count FROM patients p
                LEFT JOIN medical_history mh ON p.id = mh.patient_id
                GROUP BY p.id ORDER BY p.patient_id ASC
            ''').fetchall()
        return jsonify([dict(r) for r in rows])
    finally:
        conn.close()

@app.route('/api/patients', methods=['POST'])
def create_patient():
    data = request.get_json()
    if not data or not data.get('name', '').strip():
        return jsonify({'error': 'Name required'}), 400
    internal_id = uid()
    ts = now()
    conn = get_db()
    try:
        short_id = generate_patient_id(conn)
        conn.execute('''
            INSERT INTO patients
              (id, patient_id, name, age, gender, blood_group, phone, email,
               address, emergency_contact, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (internal_id, short_id, data['name'].strip(),
              data.get('age'), data.get('gender', ''), data.get('blood_group', ''),
              data.get('phone', ''), data.get('email', ''),
              data.get('address', ''), data.get('emergency_contact', ''), ts, ts))
        conn.commit()
        # Create upload folder for this patient
        patient_upload_dir(short_id)
        result = dict(conn.execute('SELECT * FROM patients WHERE id = ?', (internal_id,)).fetchone())
        trigger_backup('patient_added')
        return jsonify(result), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/patients/<internal_id>', methods=['GET'])
def get_patient(internal_id):
    conn = get_db()
    try:
        r = conn.execute('SELECT * FROM patients WHERE id = ?', (internal_id,)).fetchone()
        return jsonify(dict(r)) if r else (jsonify({'error': 'Not found'}), 404)
    finally:
        conn.close()

@app.route('/api/patients/<internal_id>', methods=['PUT'])
def update_patient(internal_id):
    data = request.get_json()
    if not data or not data.get('name', '').strip():
        return jsonify({'error': 'Name required'}), 400
    conn = get_db()
    try:
        if not conn.execute('SELECT id FROM patients WHERE id = ?', (internal_id,)).fetchone():
            return jsonify({'error': 'Not found'}), 404
        conn.execute('''
            UPDATE patients SET name=?, age=?, gender=?, blood_group=?, phone=?,
            email=?, address=?, emergency_contact=?, updated_at=? WHERE id=?
        ''', (data['name'].strip(), data.get('age'), data.get('gender', ''),
              data.get('blood_group', ''), data.get('phone', ''), data.get('email', ''),
              data.get('address', ''), data.get('emergency_contact', ''), now(), internal_id))
        conn.commit()
        result = dict(conn.execute('SELECT * FROM patients WHERE id = ?', (internal_id,)).fetchone())
        trigger_backup('patient_updated')
        return jsonify(result)
    finally:
        conn.close()

@app.route('/api/patients/<internal_id>', methods=['DELETE'])
def delete_patient(internal_id):
    conn = get_db()
    try:
        row = conn.execute('SELECT * FROM patients WHERE id = ?', (internal_id,)).fetchone()
        if not row:
            return jsonify({'error': 'Not found'}), 404
        short_id = row['patient_id']
        # Delete upload folder for this patient
        pid_folder = os.path.join(app.config['UPLOAD_FOLDER'], short_id)
        if os.path.exists(pid_folder):
            shutil.rmtree(pid_folder)
        conn.execute('DELETE FROM patients WHERE id = ?', (internal_id,))
        conn.commit()
        trigger_backup('patient_deleted')
        return jsonify({'message': 'Deleted'})
    finally:
        conn.close()

# ──────────────────────────────────────────────
# API — MEDICAL HISTORY
# ──────────────────────────────────────────────

@app.route('/api/patients/<internal_id>/history', methods=['GET'])
def get_history(internal_id):
    conn = get_db()
    try:
        rows = conn.execute('''
            SELECT mh.*, COUNT(r.id) as report_count FROM medical_history mh
            LEFT JOIN reports r ON mh.id = r.history_id
            WHERE mh.patient_id = ?
            GROUP BY mh.id ORDER BY mh.visit_date DESC
        ''', (internal_id,)).fetchall()
        return jsonify([dict(r) for r in rows])
    finally:
        conn.close()

@app.route('/api/patients/<internal_id>/history', methods=['POST'])
def add_history(internal_id):
    data = request.get_json()
    if not data or not data.get('visit_date', '').strip():
        return jsonify({'error': 'Date required'}), 400
    conn = get_db()
    try:
        if not conn.execute('SELECT id FROM patients WHERE id = ?', (internal_id,)).fetchone():
            return jsonify({'error': 'Patient not found'}), 404
        hid = uid()
        conn.execute('''
            INSERT INTO medical_history
              (id, patient_id, visit_date, disease, diagnosis, prescription, notes, doctor_name, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (hid, internal_id, data['visit_date'].strip(), data.get('disease', ''),
              data.get('diagnosis', ''), data.get('prescription', ''),
              data.get('notes', ''), data.get('doctor_name', ''), now()))
        conn.execute('UPDATE patients SET updated_at=? WHERE id=?', (now(), internal_id))
        conn.commit()
        result = dict(conn.execute('SELECT * FROM medical_history WHERE id = ?', (hid,)).fetchone())
        trigger_backup('visit_added')
        return jsonify(result), 201
    finally:
        conn.close()

@app.route('/api/history/<hid>', methods=['PUT'])
def update_history(hid):
    data = request.get_json()
    conn = get_db()
    try:
        row = conn.execute('SELECT * FROM medical_history WHERE id = ?', (hid,)).fetchone()
        if not row:
            return jsonify({'error': 'Not found'}), 404
        conn.execute('''
            UPDATE medical_history SET visit_date=?, disease=?, diagnosis=?,
            prescription=?, notes=?, doctor_name=? WHERE id=?
        ''', (data.get('visit_date', row['visit_date']), data.get('disease', row['disease']),
              data.get('diagnosis', row['diagnosis']), data.get('prescription', row['prescription']),
              data.get('notes', row['notes']), data.get('doctor_name', row['doctor_name']), hid))
        conn.execute('UPDATE patients SET updated_at=? WHERE id=?', (now(), row['patient_id']))
        conn.commit()
        result = dict(conn.execute('SELECT * FROM medical_history WHERE id = ?', (hid,)).fetchone())
        trigger_backup('visit_updated')
        return jsonify(result)
    finally:
        conn.close()

@app.route('/api/history/<hid>', methods=['DELETE'])
def delete_history(hid):
    conn = get_db()
    try:
        row = conn.execute('SELECT * FROM medical_history WHERE id = ?', (hid,)).fetchone()
        if not row:
            return jsonify({'error': 'Not found'}), 404
        conn.execute('DELETE FROM medical_history WHERE id = ?', (hid,))
        conn.commit()
        trigger_backup('visit_deleted')
        return jsonify({'message': 'Deleted'})
    finally:
        conn.close()

# ──────────────────────────────────────────────
# API — REPORTS / FILE UPLOAD
# ──────────────────────────────────────────────

@app.route('/api/patients/<internal_id>/reports', methods=['GET'])
def get_reports(internal_id):
    conn = get_db()
    try:
        return jsonify([dict(r) for r in conn.execute(
            'SELECT * FROM reports WHERE patient_id = ? ORDER BY uploaded_at DESC',
            (internal_id,)
        ).fetchall()])
    finally:
        conn.close()

@app.route('/api/patients/<internal_id>/reports', methods=['POST'])
def upload_report(internal_id):
    conn = get_db()
    try:
        patient = conn.execute(
            'SELECT * FROM patients WHERE id = ?', (internal_id,)
        ).fetchone()
        if not patient:
            return jsonify({'error': 'Patient not found'}), 404

        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        file = request.files['file']
        if not file.filename:
            return jsonify({'error': 'No file selected'}), 400
        if not allowed_file(file.filename):
            return jsonify({'error': f'Type not allowed. Use: {", ".join(ALLOWED_EXTENSIONS)}'}), 400

        short_id = patient['patient_id']
        orig     = secure_filename(file.filename)
        ext      = orig.rsplit('.', 1)[1].lower()
        fname    = f"{uid()}.{ext}"

        # Save to /uploads/P-0001/filename.ext
        save_dir = patient_upload_dir(short_id)
        file.save(os.path.join(save_dir, fname))

        rid = uid()
        conn.execute('''
            INSERT INTO reports
              (id, patient_id, history_id, filename, original_name, file_type, report_type, description, uploaded_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (rid, internal_id, request.form.get('history_id') or None,
              fname, orig, ext,
              request.form.get('report_type', ''),
              request.form.get('description', ''), now()))
        conn.commit()
        result = dict(conn.execute('SELECT * FROM reports WHERE id = ?', (rid,)).fetchone())
        trigger_backup('file_uploaded')
        return jsonify(result), 201
    finally:
        conn.close()

@app.route('/api/reports/<rid>', methods=['DELETE'])
def delete_report(rid):
    conn = get_db()
    try:
        row = conn.execute('SELECT r.*, p.patient_id as short_id FROM reports r JOIN patients p ON p.id = r.patient_id WHERE r.id = ?', (rid,)).fetchone()
        if not row:
            return jsonify({'error': 'Not found'}), 404
        fp = os.path.join(app.config['UPLOAD_FOLDER'], row['short_id'], row['filename'])
        if os.path.exists(fp):
            os.remove(fp)
        conn.execute('DELETE FROM reports WHERE id = ?', (rid,))
        conn.commit()
        trigger_backup('file_deleted')
        return jsonify({'message': 'Deleted'})
    finally:
        conn.close()

@app.route('/uploads/<short_id>/<filename>')
def serve_file(short_id, filename):
    safe_id   = secure_filename(short_id)
    safe_file = secure_filename(filename)
    folder    = os.path.join(app.config['UPLOAD_FOLDER'], safe_id)
    fp        = os.path.join(folder, safe_file)
    if not os.path.exists(fp):
        abort(404)
    return send_from_directory(folder, safe_file)

# ──────────────────────────────────────────────
# API — SEARCH & STATS
# ──────────────────────────────────────────────

@app.route('/api/search', methods=['GET'])
def search_disease():
    kw = request.args.get('keyword', '').strip()
    if not kw:
        return jsonify([])
    conn = get_db()
    try:
        rows = conn.execute('''
            SELECT DISTINCT p.id, p.patient_id, p.name, p.age, p.gender, p.blood_group, p.phone,
                   mh.disease, mh.diagnosis, mh.visit_date, mh.doctor_name
            FROM patients p JOIN medical_history mh ON p.id = mh.patient_id
            WHERE mh.disease LIKE ? OR mh.diagnosis LIKE ? OR mh.notes LIKE ? OR mh.prescription LIKE ?
            ORDER BY p.patient_id ASC, mh.visit_date DESC
        ''', (f'%{kw}%',) * 4).fetchall()
        return jsonify([dict(r) for r in rows])
    finally:
        conn.close()

@app.route('/api/stats', methods=['GET'])
def get_stats():
    conn = get_db()
    try:
        return jsonify({
            'total_patients':  conn.execute('SELECT COUNT(*) FROM patients').fetchone()[0],
            'total_visits':    conn.execute('SELECT COUNT(*) FROM medical_history').fetchone()[0],
            'total_reports':   conn.execute('SELECT COUNT(*) FROM reports').fetchone()[0],
            'recent_patients': conn.execute(
                'SELECT COUNT(*) FROM patients WHERE created_at >= date("now", "-30 days")'
            ).fetchone()[0],
        })
    finally:
        conn.close()

# ──────────────────────────────────────────────
# ENTRY POINT
# ──────────────────────────────────────────────

def open_browser():
    import time
    time.sleep(1.5)
    webbrowser.open('http://127.0.0.1:5000')

if __name__ == '__main__':
    init_db()
    print("=" * 55)
    print("  MediRecord is starting...")
    print("=" * 55)
    print(f"  Data folder : {BASE_DIR}")
    print(f"  Database    : {os.path.join(BASE_DIR, 'patients.db')}")
    print(f"  Uploads     : {app.config['UPLOAD_FOLDER']}")
    print("=" * 55)
    print("  Browser will open automatically.")
    print("  Keep this window open while using the app.")
    print("=" * 55)
    trigger_backup('app_startup')
    threading.Thread(target=open_browser, daemon=True).start()
    app.run(debug=False, host='127.0.0.1', port=5000, use_reloader=False)
